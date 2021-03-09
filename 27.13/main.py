import cgi
import openpyxl
from string import Template

OPTION_PASSENGER = "<option value=\"$name\">$name</option>"
OPTION_ROUTE = "<option value=\"$route\">$route</option>"

class Passenger:
    def __init__(self, name, from_city, to_city):
        self._name = name
        self._from_city = from_city
        self._to_city = to_city
    def route_cost(self, routes, rate):
        for r in routes:
            if r[0] == self._from_city and r[1] == self._to_city:
                return rate * r[2]
        return None
    def change_route(self, new_from, new_to):
        self._from_city = new_from
        self._to_city = new_to

class RouteCost:
    def __init__(self, passengers_data, routes_data, rate):
        self._passengers_data = passengers_data
        self._routes_data = routes_data
        self._rate = rate
        self._update_routes()

    def _update_routes(self):
        self._routes = []
        wb = openpyxl.load_workbook(self._routes_data)
        ws = wb.active
        for row in ws.rows:
            if len(row) == 3:
                self._routes.append((str(row[0].value), str(row[1].value), row[2].value))

    def _get_passenger_names(self):
        wb = openpyxl.load_workbook(self._passengers_data)
        ws = wb.active
        return [row[0].value for row in ws.rows if len(row) == 3]

    def _load_passenger(self, name): 
        wb = openpyxl.load_workbook(self._passengers_data)
        ws = wb.active
        for row in ws.rows:
            if row[0].value == name and len(row) == 3:
                return Passenger(str(row[0].value), str(row[1].value), str(row[2].value))

    def _change_passenger_route(self, name, new_from, new_to):
        wb = openpyxl.load_workbook(self._passengers_data)
        ws = wb.active
        for row in ws.rows:
            if row[0].value == name and len(row) == 3:
                row[1].value = new_from
                row[2].value = new_to
        wb.save(self._passengers_data)

    def _add_passenger(self, name, from_city, to_city):
        wb = openpyxl.load_workbook(self._passengers_data)
        ws = wb.active
        ws.append([name, from_city, to_city])
        wb.save(self._passengers_data)

    def _add_route(self, from_city, to_city, dist):
        wb = openpyxl.load_workbook(self._routes_data)
        ws = wb.active
        ws.append([from_city, to_city, dist])
        wb.save(self._routes_data)
        self._update_routes()

    def __call__(self, environ, start_response):
        path = environ.get("PATH_INFO", "").lstrip("/")
        params = {"result": "", "passengers" : "", "routes" : "", "cost" : ""}
        status = "200 OK"
        headers = [("Content-Type", "text/html; charset=utf-8")]

        if path == "":
            form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
            name = form.getfirst("passengers", "")
            route = form.getfirst("routes", "")
            route = route.split()
            change_route = form.getfirst("change_route", "")
            if len(route) > 1:
                if change_route:
                    self._change_passenger_route(name, route[0], route[-1])
                passenger = self._load_passenger(name)
                cost = passenger.route_cost(self._routes, self._rate)
                if cost != None:
                    params["cost"] = "Поточний маршрут: {}, вартість: {}".format(passenger._from_city + " - " + passenger._to_city, cost)
                else:
                    params["cost"] = "Пасажир не має зареєстрованого маршруту"

            names = self._get_passenger_names()
            for n in names:
                params["passengers"] += Template(OPTION_PASSENGER).substitute(name=n)
            for r in self._routes:
                rc = r[0] + " - " + r[1]
                params["routes"] += Template(OPTION_ROUTE).substitute(route=rc)

            html = "templates/routes.html"
            params["result"] += "Список маршрутів:<br>"
            for r in self._routes:
                if r[0] != "0" and r[1] != "0":
                    params["result"] += r[0] + " - " + r[1] + " ({})<br>".format(r[2])

        elif path == "add_route":
            html = "templates/add_route.html"
            form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
            from_city = form.getfirst("from", "")
            to_city = form.getfirst("to", "")
            dist = form.getfirst("dist", "")

            if from_city and to_city:
                self._add_route(from_city, to_city, dist)
                params["result"] = "Маршрут додано!"
            else:
                params["result"] = ""

        elif path == "add_passenger":
            html = "templates/add_passenger.html"
            form = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ)
            name = form.getfirst("name", "")
            from_city = form.getfirst("from", "")
            to_city = form.getfirst("to", "")

            if name:
                if not from_city or not to_city:
                    from_city = "0"
                    to_city = "0"
                self._add_passenger(name, from_city, to_city)
                params["result"] = "Пасажир доданий!"
            else:
                params["result"] = ""

        else:
            status = "404 NOT FOUND"
            html = "templates/error_404.html"

        start_response(status, headers)
        with open(html, encoding="utf-8") as f:
            page = Template(f.read()).substitute(params)
        return [bytes(page, encoding="utf-8")]

HOST = ""
PORT = 8000

if __name__ == '__main__':
    app = RouteCost("data/passengers.xlsx", "data/routes.xlsx", 10)
    from wsgiref.simple_server import make_server
    make_server(HOST, PORT, app).serve_forever()